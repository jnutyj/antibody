import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os

def highlight_excel_advanced(input_file, output_file, condition_config):
    """
    增强版函数：支持更多颜色和条件类型
    
    参数:
    input_file: 输入Excel文件路径
    output_file: 输出Excel文件路径
    condition_config: 条件配置，字典格式:
        {
            '列名': [
                {'condition': lambda x: 400 <= x <= 500, 'color': 'orange'},
                {'condition': lambda x: x > 500, 'color': 'red'},
                # 可以添加更多条件
            ],
            # 可以添加更多列
        }
    """
    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 文件 '{input_file}' 不存在")
        return False
    
    # 加载Excel文件
    try:
        wb = load_workbook(input_file)
        ws = wb.active
    except Exception as e:
        print(f"错误: 无法打开Excel文件 '{input_file}': {e}")
        return False
    
    # 定义颜色样式
    color_styles = {
        'orange': {
            'font': Font(color="FFA500", bold=True),
            'fill': PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        },
        'red': {
            'font': Font(color="FF0000", bold=True),
            'fill': PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        },
        'green': {
            'font': Font(color="008000", bold=True),
            'fill': PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        },
        'blue': {
            'font': Font(color="0000FF", bold=True),
            'fill': PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
        }
        # 可以添加更多颜色
    }
    
    # 获取列名和对应的列索引
    column_map = {}
    for idx, cell in enumerate(ws[1]):  # 假设第一行是列名
        column_map[cell.value] = idx + 1  # openpyxl使用1-based索引
    
    print(f"检测到的列名: {list(column_map.keys())}")
    
    # 处理每列的条件
    for col_name, conditions in condition_config.items():
        if col_name not in column_map:
            print(f"警告: 列 '{col_name}' 不存在于Excel文件中")                              
            continue
            
        col_idx = column_map[col_name]
        print(f"处理列 '{col_name}'")
        
        # 遍历该列的所有单元格（跳过标题行）
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            
            # 检查单元格是否为数字
            if not isinstance(cell.value, (int, float)):
                continue
                
            # 检查所有条件
            for condition in conditions:
                if condition['condition'](cell.value):
                    color = condition['color']
                    cell.font = color_styles[color]['font']
                    cell.fill = color_styles[color]['fill']
                    break  # 只应用第一个匹配的条件
    
    # 保存修改后的Excel文件
    try:
        wb.save(output_file)
        print(f"已保存标记后的文件: {output_file}")
        return True
    except Exception as e:
        print(f"错误: 无法保存文件 '{output_file}': {e}")
        return False

# 使用示例
if __name__ == "__main__":
    # 指定输入和输出文件路径
    input_excel = "output_Fv.xlsx"
    
    # 创建输出文件路径（在相同目录下添加"_highlighted"后缀）
    directory, filename = os.path.split(input_excel)
    name, ext = os.path.splitext(filename)
    output_excel = os.path.join(directory, f"{name}_highlighted{ext}")
    
    # 定义条件配置（使用lambda函数）
    conditions = {
        'All_AggScore': [
            {'condition': lambda x: 180 <= x <= 301, 'color': 'orange'},
            {'condition': lambda x: x > 301, 'color': 'red'}
        ],
        'All_Formal_Charge': [
            {'condition': lambda x: x > 8, 'color': 'red'}
        ],
        'All_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 899 <= x <= 1199, 'color': 'orange'},
            {'condition': lambda x: x > 1199, 'color': 'red'}
        ],
        'All_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 703 <= x <= 931, 'color': 'orange'},
            {'condition': lambda x: x > 931, 'color': 'red'}
        ],
        'All_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 460 <= x <= 720, 'color': 'orange'},
            {'condition': lambda x: x > 720, 'color': 'red'}
        ],
        'All_Negative_Patch_Energy': [
            {'condition': lambda x: 2357 <= x <= 3013, 'color': 'orange'},
            {'condition': lambda x: x > 3013, 'color': 'red'}
        ],
        'All_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 1384 <= x <= 1917, 'color': 'orange'},
            {'condition': lambda x: x > 1917, 'color': 'red'}
        ],
        'All_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 1080 <= x <= 1570, 'color': 'orange'},
            {'condition': lambda x: x > 1570, 'color': 'red'}
        ],
        'All_Positive_Patch_Energy': [
            {'condition': lambda x: 2454 <= x <= 2987, 'color': 'orange'},
            {'condition': lambda x: x > 2987, 'color': 'red'}
        ],
        'All_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 1611 <= x <= 1986, 'color': 'orange'},
            {'condition': lambda x: x > 1986, 'color': 'red'}
        ],
        'All_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 1463 <= x <= 1903, 'color': 'orange'},
            {'condition': lambda x: x > 1903, 'color': 'red'}
        ],
        'CDRH_AggScore': [
            {'condition': lambda x: 104 <= x <= 204, 'color': 'orange'},
            {'condition': lambda x: x > 204, 'color': 'red'}
        ],
        'CDRH_Formal_Charge': [
            {'condition': lambda x: x > 4, 'color': 'red'}
        ],
        'CDRH_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 341 <= x <= 506, 'color': 'orange'},
            {'condition': lambda x: x > 506, 'color': 'red'}
        ],
        'CDRH_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 303 <= x <= 463, 'color': 'orange'},
            {'condition': lambda x: x > 463, 'color': 'red'}
        ],
        'CDRH_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 247 <= x <= 454, 'color': 'orange'},
            {'condition': lambda x: x > 454, 'color': 'red'}
        ],
        'CDRH_Negative_Patch_Energy': [
            {'condition': lambda x: 527 <= x <= 968, 'color': 'orange'},
            {'condition': lambda x: x > 968, 'color': 'red'}
        ],
        'CDRH_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 349 <= x <= 752, 'color': 'orange'},
            {'condition': lambda x: x > 752, 'color': 'red'}
        ],
        'CDRH_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 281 <= x <= 668, 'color': 'orange'},
            {'condition': lambda x: x > 668, 'color': 'red'}
        ],
        'CDRH_Positive_Patch_Energy': [
            {'condition': lambda x: 497 <= x <= 951, 'color': 'orange'},
            {'condition': lambda x: x > 951, 'color': 'red'}
        ],
        'CDRH_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 344 <= x <= 840, 'color': 'orange'},
            {'condition': lambda x: x > 840, 'color': 'red'}
        ],
        'CDRH_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 307 <= x <= 774, 'color': 'orange'},
            {'condition': lambda x: x > 774, 'color': 'red'}
        ],
        'CDRL_AggScore': [
            {'condition': lambda x: 42 <= x <= 124, 'color': 'orange'},
            {'condition': lambda x: x > 124, 'color': 'red'}
        ],
        'CDRL_Formal_Charge': [
            {'condition': lambda x: x > 8, 'color': 'red'}
        ],
        'CDRL_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 172 <= x <= 340, 'color': 'orange'},
            {'condition': lambda x: x > 340, 'color': 'red'}
        ],
        'CDRL_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 153 <= x <= 327, 'color': 'orange'},
            {'condition': lambda x: x > 327, 'color': 'red'}
        ],
        'CDRL_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 124 <= x <= 308, 'color': 'orange'},
            {'condition': lambda x: x > 308, 'color': 'red'}
        ],
        'CDRL_Negative_Patch_Energy': [
            {'condition': lambda x: 431 <= x <= 888, 'color': 'orange'},
            {'condition': lambda x: x > 888, 'color': 'red'}
        ],
        'CDRL_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 282 <= x <= 721, 'color': 'orange'},
            {'condition': lambda x: x > 721, 'color': 'red'}
        ],
        'CDRL_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 221 <= x <= 553, 'color': 'orange'},
            {'condition': lambda x: x > 553, 'color': 'red'}
        ],
        'CDRL_Positive_Patch_Energy': [
            {'condition': lambda x: 422 <= x <= 865, 'color': 'orange'},
            {'condition': lambda x: x > 865, 'color': 'red'}
        ],
        'CDRL_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 291 <= x <= 668, 'color': 'orange'},
            {'condition': lambda x: x > 668, 'color': 'red'}
        ],
        'CDRL_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 273 <= x <= 586, 'color': 'orange'},
            {'condition': lambda x: x > 586, 'color': 'red'}
        ],
        'CDR_AggScore': [
            {'condition': lambda x: 127 <= x <= 258, 'color': 'orange'},
            {'condition': lambda x: x > 258, 'color': 'red'}
        ],
        'CDR_Aggrescan_a4v': [
            {'condition': lambda x: 6 <= x <= 18, 'color': 'orange'},
            {'condition': lambda x: x > 18, 'color': 'red'}
        ],
        'CDR_Aggrescan_a4v_pos': [
            {'condition': lambda x: 15 <= x <= 23, 'color': 'orange'},
            {'condition': lambda x: x > 23, 'color': 'red'}
        ],
        'CDR_Formal_Charge': [
            {'condition': lambda x: x > 7, 'color': 'red'}
        ],
        'CDR_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 454 <= x <= 658, 'color': 'orange'},
            {'condition': lambda x: x > 658, 'color': 'red'}
        ],
        'CDR_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 403 <= x <= 595, 'color': 'orange'},
            {'condition': lambda x: x > 595, 'color': 'red'}
        ],
        'CDR_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 318 <= x <= 530, 'color': 'orange'},
            {'condition': lambda x: x > 530, 'color': 'red'}
        ],
        'CDR_Negative_Patch_Energy': [
            {'condition': lambda x: 852 <= x <= 1381, 'color': 'orange'},
            {'condition': lambda x: x > 1381, 'color': 'red'}
        ],
        'CDR_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 515 <= x <= 1057, 'color': 'orange'},
            {'condition': lambda x: x > 1057, 'color': 'red'}
        ],
        'CDR_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 401 <= x <= 871, 'color': 'orange'},
            {'condition': lambda x: x > 871, 'color': 'red'}
        ],
        'CDR_Positive_Patch_Energy': [
            {'condition': lambda x: 843 <= x <= 1407, 'color': 'orange'},
            {'condition': lambda x: x > 1407, 'color': 'red'}
        ],
        'CDR_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 569 <= x <= 1097, 'color': 'orange'},
            {'condition': lambda x: x > 1097, 'color': 'red'}
        ],
        'CDR_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 524 <= x <= 986, 'color': 'orange'},
            {'condition': lambda x: x > 986, 'color': 'red'}
        ],
        'FRH_AggScore': [
            {'condition': lambda x: 30 <= x <= 90, 'color': 'orange'},
            {'condition': lambda x: x > 90, 'color': 'red'}
        ],
        'FRH_Aggrescan_a4v': [
            {'condition': lambda x: 2 <= x <= 5, 'color': 'orange'},
            {'condition': lambda x: x > 5, 'color': 'red'}
        ],
        'FRH_Aggrescan_a4v_pos': [
            {'condition': lambda x: 15 <= x <= 17, 'color': 'orange'},
            {'condition': lambda x: x > 17, 'color': 'red'}
        ],
        'FRH_Formal_Charge': [
            {'condition': lambda x: x > 6, 'color': 'red'}
        ],
        'FRH_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 208 <= x <= 341, 'color': 'orange'},
            {'condition': lambda x: x > 341, 'color': 'red'}
        ],
        'FRH_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 151 <= x <= 298, 'color': 'orange'},
            {'condition': lambda x: x > 298, 'color': 'red'}
        ],
        'FRH_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 91 <= x <= 200, 'color': 'orange'},
            {'condition': lambda x: x > 200, 'color': 'red'}
        ],
        'FRH_Negative_Patch_Energy': [
            {'condition': lambda x: 631 <= x <= 994, 'color': 'orange'},
            {'condition': lambda x: x > 994, 'color': 'red'}
        ],
        'FRH_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 364 <= x <= 677, 'color': 'orange'},
            {'condition': lambda x: x > 677, 'color': 'red'}
        ],
        'FRH_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 246 <= x <= 473, 'color': 'orange'},
            {'condition': lambda x: x > 473, 'color': 'red'}
        ],
        'FRH_Positive_Patch_Energy': [
            {'condition': lambda x: 886 <= x <= 1153, 'color': 'orange'},
            {'condition': lambda x: x > 1153, 'color': 'red'}
        ],
        'FRH_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 602 <= x <= 849, 'color': 'orange'},
            {'condition': lambda x: x > 849, 'color': 'red'}
        ],
        'FRH_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 524 <= x <= 817, 'color': 'orange'},
            {'condition': lambda x: x > 817, 'color': 'red'}
        ],
        'FRL_AggScore': [
            {'condition': lambda x: 57 <= x <= 102, 'color': 'orange'},
            {'condition': lambda x: x > 102, 'color': 'red'}
        ],
        'FRL_Formal_Charge': [
            {'condition': lambda x: x > 2, 'color': 'red'}
        ],
        'FRL_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 227 <= x <= 317, 'color': 'orange'},
            {'condition': lambda x: x > 317, 'color': 'red'}
        ],
        'FRL_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 155 <= x <= 227, 'color': 'orange'},
            {'condition': lambda x: x > 227, 'color': 'red'}
        ],
        'FRL_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 66 <= x <= 170, 'color': 'orange'},
            {'condition': lambda x: x > 170, 'color': 'red'}
        ],
        'FRL_Negative_Patch_Energy': [
            {'condition': lambda x: 703 <= x <= 955, 'color': 'orange'},
            {'condition': lambda x: x > 955, 'color': 'red'}
        ],
        'FRL_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 455 <= x <= 674, 'color': 'orange'},
            {'condition': lambda x: x > 674, 'color': 'red'}
        ],
        'FRL_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 387 <= x <= 653, 'color': 'orange'},
            {'condition': lambda x: x > 653, 'color': 'red'}
        ],
        'FRL_Positive_Patch_Energy': [
            {'condition': lambda x: 727 <= x <= 1009, 'color': 'orange'},
            {'condition': lambda x: x > 1009, 'color': 'red'}
        ],
        'FRL_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 482 <= x <= 782, 'color': 'orange'},
            {'condition': lambda x: x > 782, 'color': 'red'}
        ],
        'FRL_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 437 <= x <= 666, 'color': 'orange'},
            {'condition': lambda x: x > 666, 'color': 'red'}
        ],
        'FR_AggScore': [
            {'condition': lambda x: 77 <= x <= 121, 'color': 'orange'},
            {'condition': lambda x: x > 121, 'color': 'red'}
        ],
        'FR_Aggrescan_a4v': [
            {'condition': lambda x: 2 <= x <= 8, 'color': 'orange'},
            {'condition': lambda x: x > 8, 'color': 'red'}
        ],
        'FR_Aggrescan_a4v_pos': [
            {'condition': lambda x: 28 <= x <= 33, 'color': 'orange'},
            {'condition': lambda x: x > 33, 'color': 'red'}
        ],
        'FR_Formal_Charge': [
            {'condition': lambda x: x > 7, 'color': 'red'}
        ],
        'FR_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 407 <= x <= 558, 'color': 'orange'},
            {'condition': lambda x: x > 558, 'color': 'red'}
        ],
        'FR_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 282 <= x <= 439, 'color': 'orange'},
            {'condition': lambda x: x > 439, 'color': 'red'}
        ],
        'FR_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 130 <= x <= 291, 'color': 'orange'},
            {'condition': lambda x: x > 291, 'color': 'red'}
        ],
        'FR_Negative_Patch_Energy': [
            {'condition': lambda x: 1260 <= x <= 1607, 'color': 'orange'},
            {'condition': lambda x: x > 1607, 'color': 'red'}
        ],
        'FR_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 745 <= x <= 1102, 'color': 'orange'},
            {'condition': lambda x: x > 1102, 'color': 'red'}
        ],
        'FR_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 564 <= x <= 940, 'color': 'orange'},
            {'condition': lambda x: x > 940, 'color': 'red'}
        ],
        'FR_Positive_Patch_Energy': [
            {'condition': lambda x: 1536 <= x <= 1880, 'color': 'orange'},
            {'condition': lambda x: x > 1880, 'color': 'red'}
        ],
        'FR_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 1012 <= x <= 1350, 'color': 'orange'},
            {'condition': lambda x: x > 1350, 'color': 'red'}
        ],
        'FR_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 884 <= x <= 1293, 'color': 'orange'},
            {'condition': lambda x: x > 1293, 'color': 'red'}
        ],
        'Formal_Charge_eV': [
            {'condition': lambda x: x > 8, 'color': 'red'}
        ],
        'H1_AggScore': [
            {'condition': lambda x: 12 <= x <= 55, 'color': 'orange'},
            {'condition': lambda x: x > 55, 'color': 'red'}
        ],
        'H1_Formal_Charge': [
            {'condition': lambda x: x > 2, 'color': 'red'}
        ],
        'H1_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 49 <= x <= 130, 'color': 'orange'},
            {'condition': lambda x: x > 130, 'color': 'red'}
        ],
        'H1_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 42 <= x <= 126, 'color': 'orange'},
            {'condition': lambda x: x > 126, 'color': 'red'}
        ],
        'H1_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 38 <= x <= 103, 'color': 'orange'},
            {'condition': lambda x: x > 103, 'color': 'red'}
        ],
        'H1_Negative_Patch_Energy': [
            {'condition': lambda x: 93 <= x <= 196, 'color': 'orange'},
            {'condition': lambda x: x > 196, 'color': 'red'}
        ],
        'H1_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 80 <= x <= 175, 'color': 'orange'},
            {'condition': lambda x: x > 175, 'color': 'red'}
        ],
        'H1_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 80 <= x <= 175, 'color': 'orange'},
            {'condition': lambda x: x > 175, 'color': 'red'}
        ],
        'H1_Positive_Patch_Energy': [
            {'condition': lambda x: 56 <= x <= 334, 'color': 'orange'},
            {'condition': lambda x: x > 334, 'color': 'red'}
        ],
        'H1_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 1 < x <= 305, 'color': 'orange'},
            {'condition': lambda x: x > 305, 'color': 'red'}
        ],
        'H1_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 1 < x <= 305, 'color': 'orange'},
            {'condition': lambda x: x > 305, 'color': 'red'}
        ],
        'H2_AggScore': [
            {'condition': lambda x: 42 <= x <= 118, 'color': 'orange'},
            {'condition': lambda x: x > 118, 'color': 'red'}
        ],
        'H2_Formal_Charge': [
            {'condition': lambda x: x > 4, 'color': 'red'}
        ],
        'H2_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 147 <= x <= 278, 'color': 'orange'},
            {'condition': lambda x: x > 278, 'color': 'red'}
        ],
        'H2_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 127 <= x <= 250, 'color': 'orange'},
            {'condition': lambda x: x > 250, 'color': 'red'}
        ],
        'H2_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 107 <= x <= 233, 'color': 'orange'},
            {'condition': lambda x: x > 233, 'color': 'red'}
        ],
        'H2_Negative_Patch_Energy': [
            {'condition': lambda x: 313 <= x <= 708, 'color': 'orange'},
            {'condition': lambda x: x > 708, 'color': 'red'}
        ],
        'H2_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 221 <= x <= 643, 'color': 'orange'},
            {'condition': lambda x: x > 643, 'color': 'red'}
        ],
        'H2_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 184 <= x <= 585, 'color': 'orange'},
            {'condition': lambda x: x > 585, 'color': 'red'}
        ],
        'H2_Positive_Patch_Energy': [
            {'condition': lambda x: 342 <= x <= 545, 'color': 'orange'},
            {'condition': lambda x: x > 545, 'color': 'red'}
        ],
        'H2_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 257 <= x <= 500, 'color': 'orange'},
            {'condition': lambda x: x > 500, 'color': 'red'}
        ],
        'H2_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 246 <= x <= 457, 'color': 'orange'},
            {'condition': lambda x: x > 457, 'color': 'red'}
        ],
        'H3_AggScore': [
            {'condition': lambda x: 85 <= x <= 193, 'color': 'orange'},
            {'condition': lambda x: x > 193, 'color': 'red'}
        ],
        'H3_Formal_Charge': [
            {'condition': lambda x: x > 2, 'color': 'red'}
        ],
        'H3_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 228 <= x <= 408, 'color': 'orange'},
            {'condition': lambda x: x > 408, 'color': 'red'}
        ],
        'H3_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 209 <= x <= 400, 'color': 'orange'},
            {'condition': lambda x: x > 400, 'color': 'red'}
        ],
        'H3_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 178 <= x <= 400, 'color': 'orange'},
            {'condition': lambda x: x > 400, 'color': 'red'}
        ],
        'H3_Negative_Patch_Energy': [
            {'condition': lambda x: 226 <= x <= 592, 'color': 'orange'},
            {'condition': lambda x: x > 592, 'color': 'red'}
        ],
        'H3_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 155 <= x <= 384, 'color': 'orange'},
            {'condition': lambda x: x > 384, 'color': 'red'}
        ],
        'H3_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 121 <= x <= 342, 'color': 'orange'},
            {'condition': lambda x: x > 342, 'color': 'red'}
        ],
        'H3_Positive_Patch_Energy': [
            {'condition': lambda x: 174 <= x <= 513, 'color': 'orange'},
            {'condition': lambda x: x > 513, 'color': 'red'}
        ],
        'H3_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 129 <= x <= 442, 'color': 'orange'},
            {'condition': lambda x: x > 442, 'color': 'red'}
        ],
        'H3_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 122 <= x <= 390, 'color': 'orange'},
            {'condition': lambda x: x > 390, 'color': 'red'}
        ],
        'HFR1_AggScore': [
            {'condition': lambda x: 4 <= x <= 71, 'color': 'orange'},
            {'condition': lambda x: x > 71, 'color': 'red'}
        ],
        'HFR1_Formal_Charge': [
            {'condition': lambda x: x > 4, 'color': 'red'}
        ],
        'HFR1_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 117 <= x <= 247, 'color': 'orange'},
            {'condition': lambda x: x > 247, 'color': 'red'}
        ],
        'HFR1_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 95 <= x <= 223, 'color': 'orange'},
            {'condition': lambda x: x > 223, 'color': 'red'}
        ],
        'HFR1_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 86 <= x <= 200, 'color': 'orange'},
            {'condition': lambda x: x > 200, 'color': 'red'}
        ],
        'HFR1_Negative_Patch_Energy': [
            {'condition': lambda x: 343 <= x <= 531, 'color': 'orange'},
            {'condition': lambda x: x > 531, 'color': 'red'}
        ],
        'HFR1_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 213 <= x <= 434, 'color': 'orange'},
            {'condition': lambda x: x > 434, 'color': 'red'}
        ],
        'HFR1_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 188 <= x <= 364, 'color': 'orange'},
            {'condition': lambda x: x > 364, 'color': 'red'}
        ],
        'HFR1_Positive_Patch_Energy': [
            {'condition': lambda x: 488 <= x <= 708, 'color': 'orange'},
            {'condition': lambda x: x > 708, 'color': 'red'}
        ],
        'HFR1_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 346 <= x <= 528, 'color': 'orange'},
            {'condition': lambda x: x > 528, 'color': 'red'}
        ],
        'HFR1_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 292 <= x <= 528, 'color': 'orange'},
            {'condition': lambda x: x > 528, 'color': 'red'}
        ],
        'HFR2_AggScore': [
            {'condition': lambda x: 10 <= x <= 26, 'color': 'orange'},
            {'condition': lambda x: x > 26, 'color': 'red'}
        ],
        'HFR2_Formal_Charge': [
            {'condition': lambda x: x > 3, 'color': 'red'}
        ],
        'HFR2_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 44 <= x <= 104, 'color': 'orange'},
            {'condition': lambda x: x > 104, 'color': 'red'}
        ],
        'HFR2_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 21 <= x <= 98, 'color': 'orange'},
            {'condition': lambda x: x > 98, 'color': 'red'}
        ],
        'HFR2_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 1 < x <= 98, 'color': 'orange'},
            {'condition': lambda x: x > 98, 'color': 'red'}
        ],
        'HFR2_Negative_Patch_Energy': [
            {'condition': lambda x: 91 <= x <= 226, 'color': 'orange'},
            {'condition': lambda x: x > 226, 'color': 'red'}
        ],
        'HFR2_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 52 <= x <= 178, 'color': 'orange'},
            {'condition': lambda x: x > 178, 'color': 'red'}
        ],
        'HFR2_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 52 <= x <= 138, 'color': 'orange'},
            {'condition': lambda x: x > 138, 'color': 'red'}
        ],
        'HFR2_Positive_Patch_Energy': [
            {'condition': lambda x: 165 <= x <= 367, 'color': 'orange'},
            {'condition': lambda x: x > 367, 'color': 'red'}
        ],
        'HFR2_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 129 <= x <= 322, 'color': 'orange'},
            {'condition': lambda x: x > 322, 'color': 'red'}
        ],
        'HFR2_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 129 <= x <= 282, 'color': 'orange'},
            {'condition': lambda x: x > 282, 'color': 'red'}
        ],
        'HFR3_AggScore': [
            {'condition': lambda x: 17 <= x <= 41, 'color': 'orange'},
            {'condition': lambda x: x > 41, 'color': 'red'}
        ],
        'HFR3_Formal_Charge': [
            {'condition': lambda x: x > 4, 'color': 'red'}
        ],
        'HFR3_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 69 <= x <= 170, 'color': 'orange'},
            {'condition': lambda x: x > 170, 'color': 'red'}
        ],
        'HFR3_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 54 <= x <= 140, 'color': 'orange'},
            {'condition': lambda x: x > 140, 'color': 'red'}
        ],
        'HFR3_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 1 < x <= 121, 'color': 'orange'},
            {'condition': lambda x: x > 121, 'color': 'red'}
        ],
        'HFR3_Negative_Patch_Energy': [
            {'condition': lambda x: 285 <= x <= 494, 'color': 'orange'},
            {'condition': lambda x: x > 494, 'color': 'red'}
        ],
        'HFR3_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 183 <= x <= 382, 'color': 'orange'},
            {'condition': lambda x: x > 382, 'color': 'red'}
        ],
        'HFR3_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 115 <= x <= 249, 'color': 'orange'},
            {'condition': lambda x: x > 249, 'color': 'red'}
        ],
        'HFR3_Positive_Patch_Energy': [
            {'condition': lambda x: 363 <= x <= 539, 'color': 'orange'},
            {'condition': lambda x: x > 539, 'color': 'red'}
        ],
        'HFR3_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 261 <= x <= 426, 'color': 'orange'},
            {'condition': lambda x: x > 426, 'color': 'red'}
        ],
        'HFR3_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 195 <= x <= 387, 'color': 'orange'},
            {'condition': lambda x: x > 387, 'color': 'red'}
        ],
        'HFR4_AggScore': [
            {'condition': lambda x: 7 <= x <= 29, 'color': 'orange'},
            {'condition': lambda x: x > 29, 'color': 'red'}
        ],
        'HFR4_Formal_Charge': [
            {'condition': lambda x: x > 0, 'color': 'red'}
        ],
        'HFR4_Greasy_SASA': [
            {'condition': lambda x: x > 0, 'color': 'red'}
        ],
        'HFR4_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 49 <= x <= 90, 'color': 'orange'},
            {'condition': lambda x: x > 90, 'color': 'red'}
        ],
        'HFR4_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 41 <= x <= 84, 'color': 'orange'},
            {'condition': lambda x: x > 84, 'color': 'red'}
        ],
        'HFR4_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 41 <= x <= 84, 'color': 'orange'},
            {'condition': lambda x: x > 84, 'color': 'red'}
        ],
        'HFR4_Negative_Patch_Energy': [
            {'condition': lambda x: 159 <= x <= 237, 'color': 'orange'},
            {'condition': lambda x: x > 237, 'color': 'red'}
        ],
        'HFR4_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 112 <= x <= 191, 'color': 'orange'},
            {'condition': lambda x: x > 191, 'color': 'red'}
        ],
        'HFR4_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 99 <= x <= 191, 'color': 'orange'},
            {'condition': lambda x: x > 191, 'color': 'red'}
        ],
        'HFR4_Positive_Patch_Energy': [
            {'condition': lambda x: 40 <= x <= 174, 'color': 'orange'},
            {'condition': lambda x: x > 174, 'color': 'red'}
        ],
        'HFR4_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 1 < x <= 147, 'color': 'orange'},
            {'condition': lambda x: x > 147, 'color': 'red'}
        ],
        'HFR4_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 1 < x <= 147, 'color': 'orange'},
            {'condition': lambda x: x > 147, 'color': 'red'}
        ],
        'L1_AggScore': [
            {'condition': lambda x: 18 <= x <= 79, 'color': 'orange'},
            {'condition': lambda x: x > 79, 'color': 'red'}
        ],
        'L1_Formal_Charge': [
            {'condition': lambda x: x > 4, 'color': 'red'}
        ],
        'L1_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 87 <= x <= 174, 'color': 'orange'},
            {'condition': lambda x: x > 174, 'color': 'red'}
        ],
        'L1_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 81 <= x <= 172, 'color': 'orange'},
            {'condition': lambda x: x > 172, 'color': 'red'}
        ],
        'L1_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 72 <= x <= 172, 'color': 'orange'},
            {'condition': lambda x: x > 172, 'color': 'red'}
        ],
        'L1_Negative_Patch_Energy': [
            {'condition': lambda x: 201 <= x <= 515, 'color': 'orange'},
            {'condition': lambda x: x > 515, 'color': 'red'}
        ],
        'L1_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 128 <= x <= 443, 'color': 'orange'},
            {'condition': lambda x: x > 443, 'color': 'red'}
        ],
        'L1_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 103 <= x <= 408, 'color': 'orange'},
            {'condition': lambda x: x > 408, 'color': 'red'}
        ],
        'L1_Positive_Patch_Energy': [
            {'condition': lambda x: 269 <= x <= 578, 'color': 'orange'},
            {'condition': lambda x: x > 578, 'color': 'red'}
        ],
        'L1_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 189 <= x <= 449, 'color': 'orange'},
            {'condition': lambda x: x > 449, 'color': 'red'}
        ],
        'L1_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 184 <= x <= 444, 'color': 'orange'},
            {'condition': lambda x: x > 444, 'color': 'red'}
        ],
        'L2_AggScore': [
            {'condition': lambda x: 9 <= x <= 42, 'color': 'orange'},
            {'condition': lambda x: x > 42, 'color': 'red'}
        ],
        'L2_Formal_Charge': [
            {'condition': lambda x: x > 3, 'color': 'red'}
        ],
        'L2_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 59 <= x <= 165, 'color': 'orange'},
            {'condition': lambda x: x > 165, 'color': 'red'}
        ],
        'L2_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 52 <= x <= 165, 'color': 'orange'},
            {'condition': lambda x: x > 165, 'color': 'red'}
        ],
        'L2_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 47 <= x <= 147, 'color': 'orange'},
            {'condition': lambda x: x > 147, 'color': 'red'}
        ],
        'L2_Negative_Patch_Energy': [
            {'condition': lambda x: 151 <= x <= 277, 'color': 'orange'},
            {'condition': lambda x: x > 277, 'color': 'red'}
        ],
        'L2_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 101 <= x <= 247, 'color': 'orange'},
            {'condition': lambda x: x > 247, 'color': 'red'}
        ],
        'L2_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 85 <= x <= 197, 'color': 'orange'},
            {'condition': lambda x: x > 197, 'color': 'red'}
        ],
        'L2_Positive_Patch_Energy': [
            {'condition': lambda x: 168 <= x <= 382, 'color': 'orange'},
            {'condition': lambda x: x > 382, 'color': 'red'}
        ],
        'L2_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 137 <= x <= 382, 'color': 'orange'},
            {'condition': lambda x: x > 382, 'color': 'red'}
        ],
        'L2_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 128 <= x <= 382, 'color': 'orange'},
            {'condition': lambda x: x > 382, 'color': 'red'}
        ],
        'L3_AggScore': [
            {'condition': lambda x: 28 <= x <= 95, 'color': 'orange'},
            {'condition': lambda x: x > 95, 'color': 'red'}
        ],
        'L3_Formal_Charge': [
            {'condition': lambda x: x > 2, 'color': 'red'}
        ],
        'L3_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 91 <= x <= 191, 'color': 'orange'},
            {'condition': lambda x: x > 191, 'color': 'red'}
        ],
        'L3_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 79 <= x <= 162, 'color': 'orange'},
            {'condition': lambda x: x > 162, 'color': 'red'}
        ],
        'L3_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 67 <= x <= 147, 'color': 'orange'},
            {'condition': lambda x: x > 147, 'color': 'red'}
        ],
        'L3_Negative_Patch_Energy': [
            {'condition': lambda x: 145 <= x <= 361, 'color': 'orange'},
            {'condition': lambda x: x > 361, 'color': 'red'}
        ],
        'L3_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 99 <= x <= 318, 'color': 'orange'},
            {'condition': lambda x: x > 318, 'color': 'red'}
        ],
        'L3_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 82 <= x <= 287, 'color': 'orange'},
            {'condition': lambda x: x > 287, 'color': 'red'}
        ],
        'L3_Positive_Patch_Energy': [
            {'condition': lambda x: 90 <= x <= 271, 'color': 'orange'},
            {'condition': lambda x: x > 271, 'color': 'red'}
        ],
        'L3_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 54 <= x <= 252, 'color': 'orange'},
            {'condition': lambda x: x > 252, 'color': 'red'}
        ],
        'L3_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 51 <= x <= 252, 'color': 'orange'},
            {'condition': lambda x: x > 252, 'color': 'red'}
        ],
        'LFR1_AggScore': [
            {'condition': lambda x: 32 <= x <= 79, 'color': 'orange'},
            {'condition': lambda x: x > 79, 'color': 'red'}
        ],
        'LFR1_Formal_Charge': [
            {'condition': lambda x: x > 2, 'color': 'red'}
        ],
        'LFR1_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 109 <= x <= 181, 'color': 'orange'},
            {'condition': lambda x: x > 181, 'color': 'red'}
        ],
        'LFR1_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 79 <= x <= 167, 'color': 'orange'},
            {'condition': lambda x: x > 167, 'color': 'red'}
        ],
        'LFR1_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 43 <= x <= 136, 'color': 'orange'},
            {'condition': lambda x: x > 136, 'color': 'red'}
        ],
        'LFR1_Negative_Patch_Energy': [
            {'condition': lambda x: 331 <= x <= 496, 'color': 'orange'},
            {'condition': lambda x: x > 496, 'color': 'red'}
        ],
        'LFR1_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 211 <= x <= 430, 'color': 'orange'},
            {'condition': lambda x: x > 430, 'color': 'red'}
        ],
        'LFR1_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 204 <= x <= 430, 'color': 'orange'},
            {'condition': lambda x: x > 430, 'color': 'red'}
        ],
        'LFR1_Positive_Patch_Energy': [
            {'condition': lambda x: 197 <= x <= 342, 'color': 'orange'},
            {'condition': lambda x: x > 342, 'color': 'red'}
        ],
        'LFR1_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 114 <= x <= 273, 'color': 'orange'},
            {'condition': lambda x: x > 273, 'color': 'red'}
        ],
        'LFR1_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 112 <= x <= 273, 'color': 'orange'},
            {'condition': lambda x: x > 273, 'color': 'red'}
        ],
        'LFR2_AggScore': [
            {'condition': lambda x: 22 <= x <= 48, 'color': 'orange'},
            {'condition': lambda x: x > 48, 'color': 'red'}
        ],
        'LFR2_Formal_Charge': [
            {'condition': lambda x: x > 4, 'color': 'red'}
        ],
        'LFR2_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 95 <= x <= 123, 'color': 'orange'},
            {'condition': lambda x: x > 123, 'color': 'red'}
        ],
        'LFR2_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 58 <= x <= 102, 'color': 'orange'},
            {'condition': lambda x: x > 102, 'color': 'red'}
        ],
        'LFR2_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 33 <= x <= 95, 'color': 'orange'},
            {'condition': lambda x: x > 95, 'color': 'red'}
        ],
        'LFR2_Negative_Patch_Energy': [
            {'condition': lambda x: 70 <= x <= 214, 'color': 'orange'},
            {'condition': lambda x: x > 214, 'color': 'red'}
        ],
        'LFR2_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 1 < x <= 151, 'color': 'orange'},
            {'condition': lambda x: x > 151, 'color': 'red'}
        ],
        'LFR2_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 1 < x <= 134, 'color': 'orange'},
            {'condition': lambda x: x > 134, 'color': 'red'}
        ],
        'LFR2_Positive_Patch_Energy': [
            {'condition': lambda x: 367 <= x <= 486, 'color': 'orange'},
            {'condition': lambda x: x > 486, 'color': 'red'}
        ],
        'LFR2_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 322 <= x <= 454, 'color': 'orange'},
            {'condition': lambda x: x > 454, 'color': 'red'}
        ],
        'LFR2_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 3221 <= x <= 454, 'color': 'orange'},
            {'condition': lambda x: x > 454, 'color': 'red'}
        ],
        'LFR3_AggScore': [
            {'condition': lambda x: 14 <= x <= 32, 'color': 'orange'},
            {'condition': lambda x: x > 32, 'color': 'red'}
        ],
        'LFR3_Formal_Charge': [
            {'condition': lambda x: x > 0, 'color': 'red'}
        ],
        'LFR3_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 66 <= x <= 150, 'color': 'orange'},
            {'condition': lambda x: x > 150, 'color': 'red'}
        ],
        'LFR3_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 53 <= x <= 147, 'color': 'orange'},
            {'condition': lambda x: x > 147, 'color': 'red'}
        ],
        'LFR3_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 33 <= x <= 126, 'color': 'orange'},
            {'condition': lambda x: x > 126, 'color': 'red'}
        ],
        'LFR3_Negative_Patch_Energy': [
            {'condition': lambda x: 358 <= x <= 548, 'color': 'orange'},
            {'condition': lambda x: x > 548, 'color': 'red'}
        ],
        'LFR3_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 270 <= x <= 388, 'color': 'orange'},
            {'condition': lambda x: x > 388, 'color': 'red'}
        ],
        'LFR3_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 231 <= x <= 343, 'color': 'orange'},
            {'condition': lambda x: x > 343, 'color': 'red'}
        ],
        'LFR3_Positive_Patch_Energy': [
            {'condition': lambda x: 285 <= x <= 463, 'color': 'orange'},
            {'condition': lambda x: x > 463, 'color': 'red'}
        ],
        'LFR3_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 169 <= x <= 328, 'color': 'orange'},
            {'condition': lambda x: x > 328, 'color': 'red'}
        ],
        'LFR3_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 146 <= x <= 252, 'color': 'orange'},
            {'condition': lambda x: x > 252, 'color': 'red'}
        ],
        'LFR4_AggScore': [
            {'condition': lambda x: 0.8 <= x <= 16.3, 'color': 'orange'},
            {'condition': lambda x: x > 16.3, 'color': 'red'}
        ],
        'LFR4_Formal_Charge': [
            {'condition': lambda x: x > 1, 'color': 'red'}
        ],
        'LFR4_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 53 <= x <= 93, 'color': 'orange'},
            {'condition': lambda x: x > 93, 'color': 'red'}
        ],
        'LFR4_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 48 <= x <= 88, 'color': 'orange'},
            {'condition': lambda x: x > 88, 'color': 'red'}
        ],
        'LFR4_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 48 <= x <= 88, 'color': 'orange'},
            {'condition': lambda x: x > 88, 'color': 'red'}
        ],
        'LFR4_Negative_Patch_Energy': [
            {'condition': lambda x: 210 <= x <= 384, 'color': 'orange'},
            {'condition': lambda x: x > 384, 'color': 'red'}
        ],
        'LFR4_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 154 <= x <= 336, 'color': 'orange'},
            {'condition': lambda x: x > 336, 'color': 'red'}
        ],
        'LFR4_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 154 <= x <= 336, 'color': 'orange'},
            {'condition': lambda x: x > 336, 'color': 'red'}
        ],
        'LFR4_Positive_Patch_Energy': [
            {'condition': lambda x: 188 <= x <= 238, 'color': 'orange'},
            {'condition': lambda x: x > 238, 'color': 'red'}
        ],
        'LFR4_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 177 <= x <= 218, 'color': 'orange'},
            {'condition': lambda x: x > 218, 'color': 'red'}
        ],
        'LFR4_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 177 <= x <= 218, 'color': 'orange'},
            {'condition': lambda x: x > 218, 'color': 'red'}
        ],
        'Max_Size_Hyd_Patches': [
            {'condition': lambda x: 598 <= x <= 975, 'color': 'orange'},
            {'condition': lambda x: x > 975, 'color': 'red'}
        ],
        'Max_Size_Neg_Patches': [
            {'condition': lambda x: 548 <= x <= 1265, 'color': 'orange'},
            {'condition': lambda x: x > 1265, 'color': 'red'}
        ],
        'Max_Size_Pos_Patches': [
            {'condition': lambda x: 750 <= x <= 1668, 'color': 'orange'},
            {'condition': lambda x: x > 1668, 'color': 'red'}
        ],
        'Sum_Size_Hyd_Patches': [
            {'condition': lambda x: 1500 <= x <= 1967, 'color': 'orange'},
            {'condition': lambda x: x > 1967, 'color': 'red'}
        ],
        'Sum_Size_Neg_Patches': [
            {'condition': lambda x: 2096 <= x <= 2539, 'color': 'orange'},
            {'condition': lambda x: x > 2539, 'color': 'red'}
        ],
        'Sum_Size_Pos_Patches': [
            {'condition': lambda x: 2449 <= x <= 2787, 'color': 'orange'},
            {'condition': lambda x: x > 2787, 'color': 'red'}
        ],
        'VH_AggScore': [
            {'condition': lambda x: 132 <= x <= 224, 'color': 'orange'},
            {'condition': lambda x: x > 224, 'color': 'red'}
        ],
        'VH_Fv_Formal_Charge': [
            {'condition': lambda x: x > 6, 'color': 'red'}
        ],
        'VH_Fv_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 545 <= x <= 725, 'color': 'orange'},
            {'condition': lambda x: x > 725, 'color': 'red'}
        ],
        'VH_Fv_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 450 <= x <= 580, 'color': 'orange'},
            {'condition': lambda x: x > 580, 'color': 'red'}
        ],
        'VH_Fv_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 323 <= x <= 497, 'color': 'orange'},
            {'condition': lambda x: x > 497, 'color': 'red'}
        ],
        'VH_Fv_Negative_Patch_Energy': [
            {'condition': lambda x: 1208 <= x <= 1716, 'color': 'orange'},
            {'condition': lambda x: x > 1716, 'color': 'red'}
        ],
        'VH_Fv_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 728 <= x <= 1230, 'color': 'orange'},
            {'condition': lambda x: x > 1230, 'color': 'red'}
        ],
        'VH_Fv_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 527 <= x <= 903, 'color': 'orange'},
            {'condition': lambda x: x > 903, 'color': 'red'}
        ],
        'VH_Fv_Positive_Patch_Energy': [
            {'condition': lambda x: 1331 <= x <= 1711, 'color': 'orange'},
            {'condition': lambda x: x > 1711, 'color': 'red'}
        ],
        'VH_Fv_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 860 <= x <= 1320, 'color': 'orange'},
            {'condition': lambda x: x > 1320, 'color': 'red'}
        ],
        'VH_Fv_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 771 <= x <= 1186, 'color': 'orange'},
            {'condition': lambda x: x > 1186, 'color': 'red'}
        ],
        'VL_AggScore': [
            {'condition': lambda x: 80 <= x <= 147, 'color': 'orange'},
            {'condition': lambda x: x > 147, 'color': 'red'}
        ],
        'VL_Formal_Charge': [
            {'condition': lambda x: x > 6, 'color': 'red'}
        ],
        'VL_Fv_AggScore': [
            {'condition': lambda x: 80 <= x <= 147, 'color': 'orange'},
            {'condition': lambda x: x > 147, 'color': 'red'}
        ],
        'VL_Fv_Formal_Charge': [
            {'condition': lambda x: x > 6, 'color': 'red'}
        ],
        'VL_Fv_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 413 <= x <= 677, 'color': 'orange'},
            {'condition': lambda x: x > 677, 'color': 'red'}
        ],
        'VL_Fv_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 308 <= x <= 542, 'color': 'orange'},
            {'condition': lambda x: x > 542, 'color': 'red'}
        ],
        'VL_Fv_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 187 <= x <= 445, 'color': 'orange'},
            {'condition': lambda x: x > 445, 'color': 'red'}
        ],
        'VL_Fv_Negative_Patch_Energy': [
            {'condition': lambda x: 1250 <= x <= 1781, 'color': 'orange'},
            {'condition': lambda x: x > 1781, 'color': 'red'}
        ],
        'VL_Fv_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 752 <= x <= 1222, 'color': 'orange'},
            {'condition': lambda x: x > 1222, 'color': 'red'}
        ],
        'VL_Fv_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 639 <= x <= 1097, 'color': 'orange'},
            {'condition': lambda x: x > 1097, 'color': 'red'}
        ],
        'VL_Fv_Positive_Patch_Energy': [
            {'condition': lambda x: 1219 <= x <= 1588, 'color': 'orange'},
            {'condition': lambda x: x > 1588, 'color': 'red'}
        ],
        'VL_Fv_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 833 <= x <= 1134, 'color': 'orange'},
            {'condition': lambda x: x > 1134, 'color': 'red'}
        ],
        'VL_Fv_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 782 <= x <= 1067, 'color': 'orange'},
            {'condition': lambda x: x > 1067, 'color': 'red'}
        ],
        'pI_PROPKA_based': [
            {'condition': lambda x: 9.62 <= x <= 10.27, 'color': 'orange'},
            {'condition': lambda x: x > 10.27, 'color': 'red'}
        ],
        'pI_model_pKa_based': [
            {'condition': lambda x: 9.17 <= x <= 9.63, 'color': 'orange'},
            {'condition': lambda x: x > 9.63, 'color': 'red'}
        ] # 可以添加更多列和条件
    }
    
    # 运行处理函数
    success = highlight_excel_advanced(input_excel, output_excel, conditions)
    
    if success:
        # 显示处理后的数据
        try:
            df = pd.read_excel(output_excel)
            print("\n处理后的数据预览:")
            print(df)
        except Exception as e:
            print(f"无法读取处理后的文件: {e}")
