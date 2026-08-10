import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import os

def highlight_excel_advanced(input_file, output_file, condition_config):
    """
    增强版函数：支持更多颜色和条件类型
    
    参数:
    input_file: 输入Excel文件路径
    output_file: 输出Excel文件路径
    condition_config: 条件配置，字典格式:
        {
            '列名': [
                {'condition': lambda x: 400 <= x <= 500, 'color': 'orange'},
                {'condition': lambda x: x > 500, 'color': 'red'},
                # 可以添加更多条件
            ],
            # 可以添加更多列
        }
    """
    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 文件 '{input_file}' 不存在")
        return False
    
    # 加载Excel文件
    try:
        wb = load_workbook(input_file)
        ws = wb.active
    except Exception as e:
        print(f"错误: 无法打开Excel文件 '{input_file}': {e}")
        return False
    
    # 定义颜色样式
    color_styles = {
        'orange': {
            'font': Font(color="FFA500", bold=True),
            'fill': PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        },
        'red': {
            'font': Font(color="FF0000", bold=True),
            'fill': PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        },
        'green': {
            'font': Font(color="008000", bold=True),
            'fill': PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        },
        'blue': {
            'font': Font(color="0000FF", bold=True),
            'fill': PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
        }
        # 可以添加更多颜色
    }
    
    # 获取列名和对应的列索引
    column_map = {}
    for idx, cell in enumerate(ws[1]):  # 假设第一行是列名
        column_map[cell.value] = idx + 1  # openpyxl使用1-based索引
    
    print(f"检测到的列名: {list(column_map.keys())}")
    
    # 处理每列的条件
    for col_name, conditions in condition_config.items():
        if col_name not in column_map:
            print(f"警告: 列 '{col_name}' 不存在于Excel文件中")                              
            continue
            
        col_idx = column_map[col_name]
        print(f"处理列 '{col_name}'")
        
        # 遍历该列的所有单元格（跳过标题行）
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            
            # 检查单元格是否为数字
            if not isinstance(cell.value, (int, float)):
                continue
                
            # 检查所有条件
            for condition in conditions:
                if condition['condition'](cell.value):
                    color = condition['color']
                    cell.font = color_styles[color]['font']
                    cell.fill = color_styles[color]['fill']
                    break  # 只应用第一个匹配的条件
    
    # 保存修改后的Excel文件
    try:
        wb.save(output_file)
        print(f"已保存标记后的文件: {output_file}")
        return True
    except Exception as e:
        print(f"错误: 无法保存文件 '{output_file}': {e}")
        return False

# 使用示例
if __name__ == "__main__":
    # 指定输入和输出文件路径
    input_excel = "output.xlsx"
    
    # 创建输出文件路径（在相同目录下添加"_highlighted"后缀）
    directory, filename = os.path.split(input_excel)
    name, ext = os.path.splitext(filename)
    output_excel = os.path.join(directory, f"{name}_highlighted{ext}")
    
    # 定义条件配置（使用lambda函数）
    conditions = {
        'All_AggScore': [
            {'condition': lambda x: 126 <= x <= 188, 'color': 'orange'},
            {'condition': lambda x: x > 188, 'color': 'red'}
        ],
        'All_Aggrescan_a4v': [
            {'condition': lambda x: 1.049 <= x <= 7.313, 'color': 'orange'},
            {'condition': lambda x: x > 7.313, 'color': 'red'}
        ],
        'All_Formal_Charge': [
            {'condition': lambda x: 3.3 <= x <= 5, 'color': 'orange'},
            {'condition': lambda x: x > 5, 'color': 'red'}
        ],
        'CDR_AggScore': [
            {'condition': lambda x: 102 <= x <= 130, 'color': 'orange'},
            {'condition': lambda x: x > 130, 'color': 'red'}
        ],
        'CDR_Aggrescan_a4v': [
            {'condition': lambda x: 4.34 <= x <= 7.45, 'color': 'orange'},
            {'condition': lambda x: x > 7.45, 'color': 'red'}
        ],
        'CDR_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 361 <= x <= 467, 'color': 'orange'},
            {'condition': lambda x: x > 467, 'color': 'red'}
        ],
        'CDR_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 326 <= x <= 411, 'color': 'orange'},
            {'condition': lambda x: x > 411, 'color': 'red'}
        ],
         'CDR_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 261 <= x <= 374, 'color': 'orange'},
            {'condition': lambda x: x > 374, 'color': 'red'}
        ],
         'CDR_Negative_Patch_Energy': [
            {'condition': lambda x: 744 <= x <= 1422, 'color': 'orange'},
            {'condition': lambda x: x > 1422, 'color': 'red'}
        ],
         'CDR_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 525 <= x <= 1193, 'color': 'orange'},
            {'condition': lambda x: x > 1193, 'color': 'red'}
        ],
         'CDR_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 490 <= x <= 1068, 'color': 'orange'},
            {'condition': lambda x: x > 1068, 'color': 'red'}
        ],
         'CDR_Positive_Patch_Energy': [
            {'condition': lambda x: 680 <= x <= 808, 'color': 'orange'},
            {'condition': lambda x: x > 808, 'color': 'red'}
        ],
         'CDR_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 489 <= x <= 576, 'color': 'orange'},
            {'condition': lambda x: x > 576, 'color': 'red'}
        ],
         'CDR_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 489 <= x <= 560, 'color': 'orange'},
            {'condition': lambda x: x > 560, 'color': 'red'}
        ],
         'CDR_Zeta_Potential': [
            {'condition': lambda x: 3.364 <= x <= 6.725, 'color': 'orange'},
            {'condition': lambda x: x > 6.725, 'color': 'red'}
        ],
         'CDR_Zyggregator_profile_smoothed': [
            {'condition': lambda x: -1 <= x <= 0, 'color': 'orange'},
            {'condition': lambda x: x > 0, 'color': 'red'}
        ],
         'CDR_Zyggregator_profile_smoothed_pos': [
            {'condition': lambda x: 10.296 <= x <= 12.927, 'color': 'orange'},
            {'condition': lambda x: x > 12.927, 'color': 'red'}
        ],
         'FR_AggScore': [
            {'condition': lambda x: 48.653 <= x <= 59.279, 'color': 'orange'},
            {'condition': lambda x: x > 59.279, 'color': 'red'}
        ],
         'FR_Aggrescan_a4v': [
            {'condition': lambda x: -1.639 <= x <= -0.230, 'color': 'orange'},
            {'condition': lambda x: x > -0.230, 'color': 'red'}
        ],
         'FR_Greasy_SASA': [
            {'condition': lambda x: 0.000001 <= x <= 0.007, 'color': 'orange'},
            {'condition': lambda x: x > 0.007, 'color': 'red'}
        ],
         'FR_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 283 <= x <= 332, 'color': 'orange'},
            {'condition': lambda x: x > 332, 'color': 'red'}
        ],
         'FR_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 233 <= x <= 282, 'color': 'orange'},
            {'condition': lambda x: x > 282, 'color': 'red'}
        ],
         'FR_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 128 <= x <= 218, 'color': 'orange'},
            {'condition': lambda x: x > 218, 'color': 'red'}
        ],
         'FR_Negative_Patch_Energy': [
            {'condition': lambda x: 815 <= x <= 991, 'color': 'orange'},
            {'condition': lambda x: x > 991, 'color': 'red'}
        ],
         'FR_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 514 <= x <= 717, 'color': 'orange'},
            {'condition': lambda x: x > 717, 'color': 'red'}
        ],
         'FR_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 446 <= x <= 586, 'color': 'orange'},
            {'condition': lambda x: x > 586, 'color': 'red'}
        ],
         'FR_Positive_Patch_Energy': [
            {'condition': lambda x: 1075 <= x <= 1208, 'color': 'orange'},
            {'condition': lambda x: x > 1208, 'color': 'red'}
        ],
         'FR_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 752 <= x <= 898, 'color': 'orange'},
            {'condition': lambda x: x > 898, 'color': 'red'}
        ],
         'FR_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 686 <= x <= 813, 'color': 'orange'},
            {'condition': lambda x: x > 813, 'color': 'red'}
        ],
         'FR_Zeta_Potential': [
            {'condition': lambda x: 10.070 <= x <= 13.431, 'color': 'orange'},
            {'condition': lambda x: x > 13.431, 'color': 'red'}
        ],
         'FR_Zyggregator_profile_smoothed': [
            {'condition': lambda x: -35.365 <= x <= -32.187, 'color': 'orange'},
            {'condition': lambda x: x > -32.187, 'color': 'red'}
        ],
         'FR_Zyggregator_profile_smoothed_pos': [
            {'condition': lambda x: 14.083 <= x <= 16.033, 'color': 'orange'},
            {'condition': lambda x: x > 16.033, 'color': 'red'}
        ],
         'H1_AggScore': [
            {'condition': lambda x: 20.122 <= x <= 28.731, 'color': 'orange'},
            {'condition': lambda x: x > 28.731, 'color': 'red'}
        ],
         'H1_Aggrescan_a4v': [
            {'condition': lambda x: 3.446 <= x <= 4.319, 'color': 'orange'},
            {'condition': lambda x: x > 4.319, 'color': 'red'}
        ],
         'H1_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 77 <= x <= 131, 'color': 'orange'},
            {'condition': lambda x: x > 131, 'color': 'red'}
        ],
         'H1_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 67 <= x <= 131, 'color': 'orange'},
            {'condition': lambda x: x > 131, 'color': 'red'}
        ],
         'H1_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 51 <= x <= 102, 'color': 'orange'},
            {'condition': lambda x: x > 102, 'color': 'red'}
        ],
         'H1_Negative_Patch_Energy': [
            {'condition': lambda x: 119 <= x <= 331, 'color': 'orange'},
            {'condition': lambda x: x > 331, 'color': 'red'}
        ],
         'H1_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 93 <= x <= 322, 'color': 'orange'},
            {'condition': lambda x: x > 322, 'color': 'red'}
        ],
         'H1_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 93 <= x <= 268, 'color': 'orange'},
            {'condition': lambda x: x > 268, 'color': 'red'}
        ],
         'H1_Positive_Patch_Energy': [
            {'condition': lambda x: 139 <= x <= 177, 'color': 'orange'},
            {'condition': lambda x: x > 177, 'color': 'red'}
        ],
         'H1_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 123 <= x <= 162, 'color': 'orange'},
            {'condition': lambda x: x > 162, 'color': 'red'}
        ],
         'H1_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 123 <= x <= 142, 'color': 'orange'},
            {'condition': lambda x: x > 142, 'color': 'red'}
        ],
         'H2_AggScore': [
            {'condition': lambda x: 29.711 <= x <= 70.095, 'color': 'orange'},
            {'condition': lambda x: x > 70.095, 'color': 'red'}
        ],
         'H2_Aggrescan_a4v': [
            {'condition': lambda x: 1.412 <= x <= 4.862, 'color': 'orange'},
            {'condition': lambda x: x > 4.862, 'color': 'red'}
        ],
         'H2_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 127 <= x <= 178, 'color': 'orange'},
            {'condition': lambda x: x > 178, 'color': 'red'}
        ],
         'H2_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 107 <= x <= 149, 'color': 'orange'},
            {'condition': lambda x: x > 149, 'color': 'red'}
        ],
         'H2_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 83 <= x <= 134, 'color': 'orange'},
            {'condition': lambda x: x > 134, 'color': 'red'}
        ],
         'H2_Negative_Patch_Energy': [
            {'condition': lambda x: 343 <= x <= 424, 'color': 'orange'},
            {'condition': lambda x: x > 424, 'color': 'red'}
        ],
         'H2_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 256 <= x <= 339, 'color': 'orange'},
            {'condition': lambda x: x > 339, 'color': 'red'}
        ],
         'H2_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 252 <= x <= 339, 'color': 'orange'},
            {'condition': lambda x: x > 339, 'color': 'red'}
        ],
         'H2_Positive_Patch_Energy': [
            {'condition': lambda x: 373 <= x <= 509, 'color': 'orange'},
            {'condition': lambda x: x > 509, 'color': 'red'}
        ],
         'H2_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 258 <= x <= 390, 'color': 'orange'},
            {'condition': lambda x: x > 390, 'color': 'red'}
        ],
         'H2_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 258 <= x <= 356, 'color': 'orange'},
            {'condition': lambda x: x > 356, 'color': 'red'}
        ],
         'H3_AggScore': [
            {'condition': lambda x: 82 <= x <= 130, 'color': 'orange'},
            {'condition': lambda x: x > 130, 'color': 'red'}
        ],
         'H3_Aggrescan_a4v': [
            {'condition': lambda x: 2.531 <= x <= 5.754, 'color': 'orange'},
            {'condition': lambda x: x > 5.754, 'color': 'red'}
        ],
         'H3_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 265 <= x <= 352, 'color': 'orange'},
            {'condition': lambda x: x > 352, 'color': 'red'}
        ],
         'H3_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 257 <= x <= 316, 'color': 'orange'},
            {'condition': lambda x: x > 316, 'color': 'red'}
        ],
         'H3_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 214 <= x <= 279, 'color': 'orange'},
            {'condition': lambda x: x > 279, 'color': 'red'}
        ],
         'H3_Negative_Patch_Energy': [
            {'condition': lambda x: 405 <= x <= 764, 'color': 'orange'},
            {'condition': lambda x: x > 764, 'color': 'red'}
        ],
         'H3_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 298 <= x <= 669, 'color': 'orange'},
            {'condition': lambda x: x > 669, 'color': 'red'}
        ],
         'H3_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 286 <= x <= 622, 'color': 'orange'},
            {'condition': lambda x: x > 622, 'color': 'red'}
        ],
         'H3_Positive_Patch_Energy': [
            {'condition': lambda x: 360 <= x <= 505, 'color': 'orange'},
            {'condition': lambda x: x > 505, 'color': 'red'}
        ],
         'H3_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 273 <= x <= 398, 'color': 'orange'},
            {'condition': lambda x: x > 398, 'color': 'red'}
        ],
         'H3_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 266 <= x <= 358, 'color': 'orange'},
            {'condition': lambda x: x > 358, 'color': 'red'}
        ],
         'HFR1_AggScore': [
            {'condition': lambda x: 11 <= x <= 29, 'color': 'orange'},
            {'condition': lambda x: x > 29, 'color': 'red'}
        ],
         'HFR1_Aggrescan_a4v': [
            {'condition': lambda x: 1.733 <= x <= 3.098, 'color': 'orange'},
            {'condition': lambda x: x > 3.098, 'color': 'red'}
        ],
         'HFR1_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 129 <= x <= 202, 'color': 'orange'},
            {'condition': lambda x: x > 202, 'color': 'red'}
        ],
         'HFR1_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 96 <= x <= 182, 'color': 'orange'},
            {'condition': lambda x: x > 182, 'color': 'red'}
        ],
         'HFR1_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 91 <= x <= 182, 'color': 'orange'},
            {'condition': lambda x: x > 182, 'color': 'red'}
        ],
         'HFR1_Negative_Patch_Energy': [
            {'condition': lambda x: 384 <= x <= 514, 'color': 'orange'},
            {'condition': lambda x: x > 514, 'color': 'red'}
        ],
         'HFR1_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 225 <= x <= 348, 'color': 'orange'},
            {'condition': lambda x: x > 348, 'color': 'red'}
        ],
         'HFR1_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 208 <= x <= 316, 'color': 'orange'},
            {'condition': lambda x: x > 316, 'color': 'red'}
        ],
         'HFR1_Positive_Patch_Energy': [
            {'condition': lambda x: 429 <= x <= 488, 'color': 'orange'},
            {'condition': lambda x: x > 488, 'color': 'red'}
        ],
         'HFR1_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 287 <= x <= 374, 'color': 'orange'},
            {'condition': lambda x: x > 374, 'color': 'red'}
        ],
         'HFR1_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 268 <= x <= 374, 'color': 'orange'},
            {'condition': lambda x: x > 374, 'color': 'red'}
        ],
         'HFR2_AggScore': [
            {'condition': lambda x: 23 <= x <= 31, 'color': 'orange'},
            {'condition': lambda x: x > 31, 'color': 'red'}
        ],
         'HFR2_Aggrescan_a4v': [
            {'condition': lambda x: -0.392 <= x <= 0.184, 'color': 'orange'},
            {'condition': lambda x: x > 0.184, 'color': 'red'}
        ],
         'HFR2_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 119 <= x <= 142, 'color': 'orange'},
            {'condition': lambda x: x > 142, 'color': 'red'}
        ],
         'HFR2_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 98 <= x <= 129, 'color': 'orange'},
            {'condition': lambda x: x > 129, 'color': 'red'}
        ],
         'HFR2_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 51 <= x <= 91, 'color': 'orange'},
            {'condition': lambda x: x > 91, 'color': 'red'}
        ],
         'HFR2_Negative_Patch_Energy': [
            {'condition': lambda x: 234 <= x <= 288, 'color': 'orange'},
            {'condition': lambda x: x > 288, 'color': 'red'}
        ],
         'HFR2_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 208 <= x <= 231, 'color': 'orange'},
            {'condition': lambda x: x > 231, 'color': 'red'}
        ],
         'HFR2_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 208 <= x <= 231, 'color': 'orange'},
            {'condition': lambda x: x > 231, 'color': 'red'}
        ],
         'HFR2_Positive_Patch_Energy': [
            {'condition': lambda x: 283 <= x <= 433, 'color': 'orange'},
            {'condition': lambda x: x > 433, 'color': 'red'}
        ],
         'HFR2_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 219 <= x <= 361, 'color': 'orange'},
            {'condition': lambda x: x > 361, 'color': 'red'}
        ],
         'HFR2_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 209 <= x <= 361, 'color': 'orange'},
            {'condition': lambda x: x > 361, 'color': 'red'}
        ],
         'HFR3_AggScoree': [
            {'condition': lambda x: 23 <= x <= 29, 'color': 'orange'},
            {'condition': lambda x: x > 29, 'color': 'red'}
        ],
         'HFR3_Aggrescan_a4v': [
            {'condition': lambda x: -1.148 <= x <= -0.406, 'color': 'orange'},
            {'condition': lambda x: x > -0.406, 'color': 'red'}
        ],
         'HFR3_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 76 <= x <= 86, 'color': 'orange'},
            {'condition': lambda x: x > 86, 'color': 'red'}
        ],
         'HFR3_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 62 <= x <= 75, 'color': 'orange'},
            {'condition': lambda x: x > 75, 'color': 'red'}
        ],
         'HFR3_Hydrophobic_Patch_Energy_gt30': [
            {'condition': lambda x: 9 <= x <= 34, 'color': 'orange'},
            {'condition': lambda x: x > 34, 'color': 'red'}
        ],
         'HFR3_Negative_Patch_Energy': [
            {'condition': lambda x: 270 <= x <= 443, 'color': 'orange'},
            {'condition': lambda x: x > 443, 'color': 'red'}
        ],
         'HFR3_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 158 <= x <= 352, 'color': 'orange'},
            {'condition': lambda x: x > 352, 'color': 'red'}
        ],
         'HFR3_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 113 <= x <= 264, 'color': 'orange'},
            {'condition': lambda x: x > 264, 'color': 'red'}
        ],
         'HFR3_Positive_Patch_Energy': [
            {'condition': lambda x: 414 <= x <= 473, 'color': 'orange'},
            {'condition': lambda x: x > 473, 'color': 'red'}
        ],
         'HFR3_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 312 <= x <= 383, 'color': 'orange'},
            {'condition': lambda x: x > 383, 'color': 'red'}
        ],
         'HFR3_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 243 <= x <= 318, 'color': 'orange'},
            {'condition': lambda x: x > 318, 'color': 'red'}
        ],
         'HFR4_AggScore': [
            {'condition': lambda x: 10 <= x <= 17, 'color': 'orange'},
            {'condition': lambda x: x > 17, 'color': 'red'}
        ],
         'HFR4_Aggrescan_a4v': [
            {'condition': lambda x: 1.942 <= x <= 2.042, 'color': 'orange'},
            {'condition': lambda x: x > 2.042, 'color': 'red'}
        ],
         'HFR4_Hydrophobic_Patch_Energy': [
            {'condition': lambda x: 88 <= x <= 113, 'color': 'orange'},
            {'condition': lambda x: x > 113, 'color': 'red'}
        ],
         'HFR4_Hydrophobic_Patch_Energy_gt15': [
            {'condition': lambda x: 81 <= x <= 94, 'color': 'orange'},
            {'condition': lambda x: x > 94, 'color': 'red'}
        ],
         'HFR4_Hydrophobic_Patch_Energy_gt30e': [
            {'condition': lambda x: 71 <= x <= 84, 'color': 'orange'},
            {'condition': lambda x: x > 84, 'color': 'red'}
        ],
         'HFR4_Negative_Patch_Energy': [
            {'condition': lambda x: 178 <= x <= 214, 'color': 'orange'},
            {'condition': lambda x: x > 214, 'color': 'red'}
        ],
         'HFR4_Negative_Patch_Energy_gt30': [
            {'condition': lambda x: 129 <= x <= 162, 'color': 'orange'},
            {'condition': lambda x: x > 162, 'color': 'red'}
        ],
         'HFR4_Negative_Patch_Energy_gt50': [
            {'condition': lambda x: 88 <= x <= 118, 'color': 'orange'},
            {'condition': lambda x: x > 118, 'color': 'red'}
        ],
         'HFR4_Positive_Patch_Energy': [
            {'condition': lambda x: 147 <= x <= 264, 'color': 'orange'},
            {'condition': lambda x: x > 264, 'color': 'red'}
        ],
         'HFR4_Positive_Patch_Energy_gt30': [
            {'condition': lambda x: 112 <= x <= 227, 'color': 'orange'},
            {'condition': lambda x: x > 227, 'color': 'red'}
        ],
         'HFR4_Positive_Patch_Energy_gt50': [
            {'condition': lambda x: 78 <= x <= 118, 'color': 'orange'},
            {'condition': lambda x: x > 118, 'color': 'red'}
        ],
         'Max_Size_Hyd_Patches': [
            {'condition': lambda x: 636 <= x <= 835, 'color': 'orange'},
            {'condition': lambda x: x > 835, 'color': 'red'}
        ],
         'Max_Size_Neg_Patches': [
            {'condition': lambda x: 486 <= x <= 1116, 'color': 'orange'},
            {'condition': lambda x: x > 1116, 'color': 'red'}
        ],
         'Max_Size_Pos_Patches': [
            {'condition': lambda x: 1172 <= x <= 1644, 'color': 'orange'},
            {'condition': lambda x: x > 1644, 'color': 'red'}
        ],
         'Sum_Size_Hyd_Patches': [
            {'condition': lambda x: 1081 <= x <= 1222, 'color': 'orange'},
            {'condition': lambda x: x > 1222, 'color': 'red'}
        ],
         'Sum_Size_Neg_Patches': [
            {'condition': lambda x: 1350 <= x <= 1858, 'color': 'orange'},
            {'condition': lambda x: x > 1858, 'color': 'red'}
        ],
         'Sum_Size_Pos_Patches': [
            {'condition': lambda x: 1701 <= x <= 1895, 'color': 'orange'},
            {'condition': lambda x: x > 1895, 'color': 'red'}
        ],
         'pI_PROPKA_based': [
            {'condition': lambda x: 10.029 <= x <= 10.463, 'color': 'orange'},
            {'condition': lambda x: x > 10.463, 'color': 'red'}
        ]
        # 可以添加更多列和条件
    }
    
    # 运行处理函数
    success = highlight_excel_advanced(input_excel, output_excel, conditions)
    
    if success:
        # 显示处理后的数据
        try:
            df = pd.read_excel(output_excel)
            print("\n处理后的数据预览:")
            print(df)
        except Exception as e:
            print(f"无法读取处理后的文件: {e}")
